import glob
import calendar
import locale
import tkinter
from tkinter import *
from tkinter.filedialog import askdirectory
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter as get_c

from alignments import *
from borders import *
from fonts import *
from patternfills import *

locale.setlocale(locale.LC_ALL, 'fr_FR')

# Constantes
# Nombre de mois dans le documents
_NB_MOIS_ = 12

# Feuille compta mensuelle
# Nb de lignes des en-têtes
_LIGN_COMPTE_ = 2
_LIGN_INTIT_ = 2
_LIGN_TITRE_ = 1
# Numéros de la première ligne de saisie
_DEB_LIGN_SAISIE_ = _LIGN_COMPTE_ + _LIGN_INTIT_ + 2
# Contenu des en-têtes
_LIB_INTIT_ = ('DATE', 'N°', 'N° CHQ', 'INTITULE')
_LIB_CAISSE_ = ('Recettes', 'Dépenses', 'Situation')
_LIB_BANQUE_ = ('Recettes', 'Dépenses', 'Situation')
_LIB_RECETTES_ = ('Recettes diverses', 'Compte à régulariser', 'Virements internes')
_LIB_DEPENSES_ = ['Virements internes', 'Epargne', 'Alimentat°', 'Produits entretien', 'Transport', 'Hygiène',
                  'Invest.', 'Santé', 'Assurances', 'Divers', 'Electricité', 'Eau', 'Impôts']
"""_LIB_DEPENSES_ = []"""
# Nombre de colonnes de chaque section du document
# Nombre de colonnes des différentes sections
_NB_COL_INTIT_ = len(_LIB_INTIT_)
_NB_COL_CAISSE_ = len(_LIB_CAISSE_)
_NB_COL_BANQUE_ = len(_LIB_BANQUE_)
_NB_COL_RECETTES_ = len(_LIB_RECETTES_)
"""_NB_COL_DEPENSES_ = len(_LIB_DEPENSES_)"""
# Numéro de colonnes de début des différentes sections
_DEB_BANQUE_ = _NB_COL_INTIT_ + _NB_COL_CAISSE_ + 1
_DEB_RECETTES_ = _NB_COL_INTIT_ + _NB_COL_CAISSE_ + _NB_COL_BANQUE_ + 1
_DEB_DEPENSES_ = _NB_COL_INTIT_ + _NB_COL_CAISSE_ + _NB_COL_BANQUE_ + _NB_COL_RECETTES_ + 1
# Nombre total de colonnes des tableaux mensuels
"""_NB_COL_TOTAL_ = _NB_COL_INTIT_ + _NB_COL_CAISSE_ + _NB_COL_BANQUE_ + _NB_COL_RECETTES_ + _NB_COL_DEPENSES_ + 1"""

# Feuille bilan
# Première ligne de saisie de la feuille bilan
_DEB_LIGN_SAISIE_B_ = 5
# Nombre de lignes séparant le tableau détaillant les dépenses mensuelles et le tableau résumé
_MARGE_INTER_TAB_BILAN_ = 2
# Numéro de la colonne sur laquelle débute la section banque
_DEB_BANQUE_BILAN_ = 2
_DEB_DEPENSES_BILAN_ = _DEB_BANQUE_BILAN_ + _NB_COL_BANQUE_
# Numéro correspondant à la colonne des totaux mensuels
"""_NUM_LAST_COL_ = _DEB_DEPENSES_BILAN_ + _NB_COL_DEPENSES_"""
# Première ligne du tableau résumé
_L_TAB_TOTAUX_ = _MARGE_INTER_TAB_BILAN_ + _DEB_LIGN_SAISIE_B_ + _NB_MOIS_
# Dernière ligne du tableau résumé
_BTM_BILAN_ = _DEB_LIGN_SAISIE_B_ + _NB_MOIS_ - 1


def genererfichiercompta():
    """
    Génère le fichier excel avec ses onglets
    :return: Fichier de comptabilité
    """
    annee = int(lannee.get())
    doc = Workbook()
    for i in range(_NB_MOIS_):
        sheet = doc.create_sheet(f"{calendar.month_name[i + 1].capitalize()} {annee}", i)
        mettreenformesheetmois(sheet)
        remplirsheetmois(doc, sheet)
    doc['Sheet'].title = f"Bilan {annee}"
    mettreenformesheetbilan(doc[f'Bilan {annee}'])
    remplirsheetbilan(doc)
    return doc


def verifnom():
    """
    Attribut un nom de fichier n'existant pas dans le repertoire choisi
    :return: Nom définitif du fichier
    """
    # Récupère la liste des fichiers présents dans le répertoire courant
    liste_fic = []
    path = lblpath['text']
    for file in glob.glob(f"{path}/*.xlsx"):
        file = file.removesuffix('.xlsx')
        liste_fic.append(file.removeprefix(f'{path}\\'))

    # Compteur pour le nom modifié
    i = 1
    ajout = ''
    nom = lenom.get()

    # Cherche si le nom de fichier existe déjà dans le répertoire à partir de la liste créée juste avant
    while True:
        nom_ok = True
        for fic in liste_fic:
            if fic == f'{nom}{ajout}':
                ajout = f'({i})'
                i += 1
                nom_ok = False
            if not nom_ok:
                break
        if nom_ok:
            break
    return f'{nom}{ajout}'


def mettreenformesheetmois(sheet):
    """
    Dimensionne une feuille de calcul de comptabilité mensuelle
    :param sheet: Feuille de calcul sur laquelle les propriétés doivent s'appliquer
    :return:
    """
    _NB_COL_DEPENSES_ = len(_LIB_DEPENSES_)
    _NB_COL_TOTAL_ = _NB_COL_INTIT_ + _NB_COL_CAISSE_ + _NB_COL_BANQUE_ + _NB_COL_RECETTES_ + _NB_COL_DEPENSES_ + 1
    nbligne = int(lignes.get()) + _DEB_LIGN_SAISIE_
    # Définition des plages de cellules à fusionner
    list_cells_a_merge = (f'A1:{sheet.cell(column=_NB_COL_INTIT_, row=_LIGN_COMPTE_).coordinate}',  # Nom de compte
                          f'{sheet.cell(column=_NB_COL_INTIT_ + 1, row=_LIGN_TITRE_).coordinate}:'
                          f'{sheet.cell(column=_NB_COL_TOTAL_, row=_LIGN_TITRE_).coordinate}',  # Titre du document
                          f'{sheet.cell(column=_NB_COL_INTIT_ + 1, row=_LIGN_TITRE_ + 1).coordinate}:'
                          f'{sheet.cell(column=_DEB_BANQUE_ - 1, row=_LIGN_TITRE_ + 1).coordinate}',  # Caisse
                          f'{sheet.cell(column=_DEB_BANQUE_, row=_LIGN_TITRE_ + 1).coordinate}:'
                          f'{sheet.cell(column=_DEB_RECETTES_ - 1, row=_LIGN_TITRE_ + 1).coordinate}',  # Banque
                          f'{sheet.cell(column=_DEB_RECETTES_, row=_LIGN_TITRE_ + 1).coordinate}:'
                          f'{sheet.cell(column=_DEB_DEPENSES_ - 1, row=_LIGN_TITRE_ + 1).coordinate}',  # Recettes
                          f'{sheet.cell(column=_DEB_DEPENSES_, row=_LIGN_TITRE_ + 1).coordinate}:'
                          f'{sheet.cell(column=_NB_COL_TOTAL_ - 1, row=_LIGN_TITRE_ + 1).coordinate}')  # Dépenses
    # Fusion des plages de cellules
    for plage in list_cells_a_merge:
        sheet.merge_cells(plage)

    # Dimensionnement des colonnes
    sheet.column_dimensions['A'].width = 8.5
    sheet.column_dimensions['B'].width = 4.73
    sheet.column_dimensions['C'].width = 8.82
    sheet.column_dimensions['D'].width = 41.5
    sheet.column_dimensions[get_c(_NB_COL_TOTAL_)].width = 10.91  # Colonne des totaux

    # Dimensionnement des lignes
    list_hauteur_ligne = (16, 15, 35, 24.5, 3.5)
    for i in range(min(len(list_hauteur_ligne), _DEB_LIGN_SAISIE_ - 1)):
        sheet.row_dimensions[i + 1].height = list_hauteur_ligne[i]
    for i in range(_DEB_LIGN_SAISIE_, nbligne):
        sheet.row_dimensions[i].height = 12.5

    # Application des couleurs de fond de cellule
    for i in range(_NB_COL_INTIT_):
        sheet[_LIGN_INTIT_ + _LIGN_COMPTE_][i].fill = fill_gris
    for i in range(_NB_COL_TOTAL_):
        sheet[_DEB_LIGN_SAISIE_ - 1][i].fill = fill_jaune

    # Définition des mises en forme conditionnelles
    cond_format_red_alert = CellIsRule(operator='lessThan', formula=[0], stopIfTrue=False, font=font_huit_red)
    cond_format_green = CellIsRule(operator='greaterThanOrEqual', formula=[0], stopIfTrue=False, font=font_huit_green)

    # Application des mises en forme conditionnelles
    liste_cell_cond_format = (f'{sheet.cell(column=_DEB_BANQUE_ - 1, row=_DEB_LIGN_SAISIE_ - 2).coordinate}',
                              f'{sheet.cell(column=_DEB_RECETTES_ - 1, row=_DEB_LIGN_SAISIE_ - 2).coordinate}')
    for cell in liste_cell_cond_format:
        sheet.conditional_formatting.add(cell, cond_format_red_alert)
        sheet.conditional_formatting.add(cell, cond_format_green)

    # Application des propriétés générales
    for row in sheet[f'A1:{get_c(_NB_COL_TOTAL_)}{nbligne - 1}']:
        for cell in row:
            cell.font = font_huit
            cell.alignment = align_base
    sheet[_LIGN_TITRE_ + 1][_NB_COL_TOTAL_ - 1].font = font_neuf

    # Propriétés de la colonne des totaux
    for i in range(_DEB_LIGN_SAISIE_, nbligne):
        sheet[f'{get_c(_NB_COL_TOTAL_)}{i}'].font = font_dix_rouge
        sheet[f'{get_c(_NB_COL_TOTAL_)}{i}'].alignment = align_droite
        sheet[f'A{i}'].alignment = align_droite

    # Propriétés de la colonne des intitulés
    for i in range(3, nbligne):
        sheet[f'{get_c(_NB_COL_INTIT_)}{i}'].font = font_dix
    sheet[f'{get_c(_NB_COL_TOTAL_)}{_DEB_LIGN_SAISIE_ - 2}'].font = font_dix

    # Propriétés des en-têtes
    for row in sheet[f'A1:{get_c(_NB_COL_TOTAL_)}{_LIGN_TITRE_ + 1}']:
        for cell in row:
            cell.font = font_douze_bold
    for row in sheet[f'A1:{get_c(_NB_COL_TOTAL_)}{_LIGN_COMPTE_ + 1}']:
        for cell in row:
            cell.alignment = align_center
    sheet[f'{get_c(_NB_COL_INTIT_)}{_DEB_LIGN_SAISIE_ + 2}'].alignment = align_titre
    sheet[f'{get_c(_NB_COL_TOTAL_)}{_DEB_LIGN_SAISIE_ - 2}'].alignment = align_titre
    cell = sheet[f'{get_c(_NB_COL_TOTAL_)}{_LIGN_COMPTE_ + 1}'].font = font_dix_bold
    cell.alignment = align_titre

    # En-têtes - Catégories
    for row in sheet[(f'{get_c(_NB_COL_INTIT_ + 1)}{_LIGN_TITRE_ + 1}:'
                      f'{get_c(_NB_COL_TOTAL_ - 1)}{_LIGN_TITRE_ + 1}')]:
        for cell in row:
            cell.font = font_huit_bold

    # Définition des zones pour les bordures
    bottom = nbligne - 1
    liste_zones = (f'{get_c(_NB_COL_TOTAL_)}{_DEB_LIGN_SAISIE_ - 3}:'
                   f'{get_c(_NB_COL_TOTAL_)}{_DEB_LIGN_SAISIE_ - 3}',  # cellule TOTAUX
                   f'{get_c(_NB_COL_INTIT_ + 1)}{_LIGN_TITRE_ + 1}:'
                   f'{get_c(_DEB_BANQUE_ - 1)}{_LIGN_TITRE_ + 1}',  # Caisse
                   f'{sheet.cell(column=_DEB_BANQUE_, row=_LIGN_TITRE_ + 1).coordinate}:'
                   f'{sheet.cell(column=_DEB_RECETTES_ - 1, row=_LIGN_TITRE_ + 1).coordinate}',  # Banque
                   f'{sheet.cell(column=_DEB_RECETTES_, row=_LIGN_TITRE_ + 1).coordinate}:'
                   f'{sheet.cell(column=_DEB_DEPENSES_ - 1, row=_LIGN_TITRE_ + 1).coordinate}',  # Recettes
                   f'{sheet.cell(column=_DEB_DEPENSES_, row=_LIGN_TITRE_ + 1).coordinate}:'
                   f'{sheet.cell(column=_NB_COL_TOTAL_ - 1, row=_LIGN_TITRE_ + 1).coordinate}',  # Dépenses
                   f'{sheet.cell(column=1, row=_DEB_LIGN_SAISIE_ - 3).coordinate}:'
                   f'{sheet.cell(column=_NB_COL_INTIT_, row=_DEB_LIGN_SAISIE_ - 3).coordinate}',  # Intitulés
                   f'{sheet.cell(column=_NB_COL_INTIT_ + 1, row=_DEB_LIGN_SAISIE_ - 3).coordinate}:'
                   f'{sheet.cell(column=_DEB_BANQUE_ - 1, row=_DEB_LIGN_SAISIE_ - 3).coordinate}',  # Liste Caisse
                   f'{sheet.cell(column=_DEB_BANQUE_, row=_DEB_LIGN_SAISIE_ - 3).coordinate}:'
                   f'{sheet.cell(column=_DEB_RECETTES_ - 1, row=_DEB_LIGN_SAISIE_ - 3).coordinate}',  # Liste Banque
                   f'{sheet.cell(column=_DEB_RECETTES_, row=_DEB_LIGN_SAISIE_ - 3).coordinate}:'
                   f'{sheet.cell(column=_DEB_DEPENSES_ - 1, row=_DEB_LIGN_SAISIE_ - 3).coordinate}',  # Liste Recettes
                   f'{sheet.cell(column=_DEB_DEPENSES_, row=_DEB_LIGN_SAISIE_ - 3).coordinate}:'
                   f'{sheet.cell(column=_NB_COL_TOTAL_ - 1, row=_DEB_LIGN_SAISIE_ - 3).coordinate}',  # Liste Dépenses
                   f'A{_LIGN_COMPTE_ + _LIGN_INTIT_}:{get_c(_NB_COL_INTIT_)}{bottom}',  # Section intitulés
                   f'{get_c(_NB_COL_INTIT_ + 1)}{_LIGN_COMPTE_ + _LIGN_INTIT_}:'
                   f'{get_c(_DEB_BANQUE_ - 1)}{bottom}',  # Section Caisse
                   f'{get_c(_DEB_BANQUE_)}{_LIGN_COMPTE_ + _LIGN_INTIT_}:'
                   f'{get_c(_DEB_RECETTES_ - 1)}{bottom}',  # Section Banque
                   f'{get_c(_DEB_RECETTES_)}{_LIGN_COMPTE_ + _LIGN_INTIT_}:'
                   f'{get_c(_DEB_DEPENSES_ - 1)}{bottom}',  # Section Recettes
                   f'{get_c(_DEB_DEPENSES_)}{_LIGN_COMPTE_ + _LIGN_INTIT_}:'
                   f'{get_c(_NB_COL_TOTAL_ - 1)}{bottom}',  # Section Dépenses
                   f'{get_c(_NB_COL_TOTAL_)}{_LIGN_COMPTE_ + _LIGN_INTIT_}:'
                   f'{get_c(_NB_COL_TOTAL_)}{bottom}')  # Section totaux

    # Application des bordures aux zones multi-cellules
    for zone in liste_zones:
        appliquerbordures(sheet[zone])

    # Formats du contenu des cellules
    monetaire_euro = '#,##0.00 €'
    date_fr = 'dd/mm/yyyy'
    # Colonne des totaux (monétaire euros deux chiffre après la virgule
    for row in sheet[(f'{get_c(_NB_COL_INTIT_ + 1)}{_LIGN_COMPTE_ + _LIGN_INTIT_}:'
                      f'{get_c(_NB_COL_TOTAL_)}{nbligne - 1}')]:
        for cell in row:
            cell.number_format = monetaire_euro
    # Colonne A (dates)
    for i in range(_DEB_LIGN_SAISIE_, nbligne):
        sheet[f'A{i}'].number_format = date_fr


def remplirsheetmois(doc, sheet):
    """
    Remplit la feuille de calcul avec le texte par défaut
    :param doc: Docuement au format .xlsx
    :param sheet: Feuille de calcul à remplir
    :return:
    """
    _NB_COL_DEPENSES_ = len(_LIB_DEPENSES_)
    _NUM_LAST_COL_ = _DEB_DEPENSES_BILAN_ + _NB_COL_DEPENSES_
    _NB_COL_TOTAL_ = _NB_COL_INTIT_ + _NB_COL_CAISSE_ + _NB_COL_BANQUE_ + _NB_COL_RECETTES_ + _NB_COL_DEPENSES_ + 1
    nbligne = int(lignes.get()) + 6
    # Remplissage des titres
    sheet['A1'].value = 'Compte chèques'
    sheet[f'{get_c(_NB_COL_INTIT_ + 1)}1'].value = 'Feuille de comptabilité'
    sheet[f'{get_c(_NB_COL_INTIT_)}{_DEB_LIGN_SAISIE_ - 2}'].value = 'totaux :'
    sheet[f'{get_c(_NB_COL_INTIT_ + 1)}{_LIGN_TITRE_ + 1}'].value = 'Caisse'
    sheet[f'{get_c(_DEB_BANQUE_)}{_LIGN_TITRE_ + 1}'].value = 'Banque'
    sheet[f'{get_c(_DEB_RECETTES_)}{_LIGN_TITRE_ + 1}'].value = 'Recettes'
    sheet[f'{get_c(_DEB_DEPENSES_)}{_LIGN_TITRE_ + 1}'].value = 'Dépenses'

    # Remplissage des champs de la ligne 3 (en-têtes)
    ligne_lib = _LIGN_TITRE_ + 2
    # Section intitulé
    for i in range(_NB_COL_INTIT_):
        sheet[f'{get_c(i + 1)}{ligne_lib}'].value = _LIB_INTIT_[i]
    # Section caisse
    for i in range(_NB_COL_CAISSE_):
        sheet[f'{get_c(i + _NB_COL_INTIT_ + 1)}{ligne_lib}'].value = _LIB_CAISSE_[i]
    # Section banque
    for i in range(_NB_COL_BANQUE_):
        sheet[f'{get_c(i + _DEB_BANQUE_)}{ligne_lib}'].value = _LIB_BANQUE_[i]
    # Section recettes
    for i in range(_NB_COL_RECETTES_):
        sheet[f'{get_c(i + _DEB_RECETTES_)}{ligne_lib}'].value = _LIB_RECETTES_[i]
    # Section dépenses
    for i in range(_NB_COL_DEPENSES_):
        sheet[f'{get_c(i + _DEB_DEPENSES_)}{ligne_lib}'].value = _LIB_DEPENSES_[i]
    # Cellule TOTAL
    sheet[f'{get_c(_NB_COL_TOTAL_)}{ligne_lib}'].value = 'TOTAL'

    # Remplissage des formules de calcul de la ligne 4 (totaux)
    for i in range(_NB_COL_INTIT_ + 1, _NB_COL_TOTAL_):
        sheet[f'{get_c(i)}{_DEB_LIGN_SAISIE_ - 2}'].value = \
            f'=SUM({get_c(i)}{_DEB_LIGN_SAISIE_}:{get_c(i)}{nbligne - 1})'
    sheet[f'{get_c(_NB_COL_INTIT_ + _NB_COL_CAISSE_)}{_DEB_LIGN_SAISIE_ - 2}'].value = \
        (f'={get_c(_NB_COL_INTIT_ + 1)}{_DEB_LIGN_SAISIE_ - 2}-'
         f'{get_c(_NB_COL_INTIT_ + 2)}{_DEB_LIGN_SAISIE_ - 2}')
    sheet[f'{get_c(_DEB_RECETTES_ - 1)}{_DEB_LIGN_SAISIE_ - 2}'].value = \
        (f'={get_c(_DEB_BANQUE_)}{_DEB_LIGN_SAISIE_ - 2}-'
         f'{get_c(_DEB_BANQUE_ + 1)}{_DEB_LIGN_SAISIE_ - 2}')

    # Remplissage des formules de calcul de la colonne des totaux
    sheet[f'{get_c(_NB_COL_TOTAL_)}{_DEB_LIGN_SAISIE_ - 2}'].value = \
        (f'=SUM({get_c(_DEB_DEPENSES_)}{_DEB_LIGN_SAISIE_ - 2}:'
         f'{get_c(_NB_COL_TOTAL_ - 1)}{_DEB_LIGN_SAISIE_ - 2})')
    for i in range(_DEB_LIGN_SAISIE_, nbligne):
        sheet[f'{get_c(_NB_COL_TOTAL_)}{i}'].value = \
            (f'={get_c(_NB_COL_INTIT_ + 1)}{i}+{get_c(_DEB_BANQUE_)}{i}-'
             f'SUM({get_c(_DEB_RECETTES_)}{i}:{get_c(_DEB_DEPENSES_ - 1)}{i})-'
             f'{get_c(_NB_COL_INTIT_ + 2)}{i}-{get_c(_DEB_BANQUE_ + 1)}{i}+'
             f'SUM({get_c(_DEB_DEPENSES_)}{i}:{get_c(_NB_COL_TOTAL_ - 1)}{i})')

    # Remplissage de la première ligne d'enregistrement
    sheet[_DEB_LIGN_SAISIE_][_NB_COL_INTIT_ - 1].value = 'Ouverture'
    month_dict = {'Janvier': '01', 'Février': '02', 'Mars': '03', 'Avril': '04', 'Mai': '05', 'Juin': '06',
                  'Juillet': '07', 'Août': '08', 'Septembre': '09', 'Octobre': '10', 'Novembre': '11', 'Décembre': '12'}
    mois = month_dict[sheet.title[:-5]]
    annee = int(sheet.title[-4:])
    sheet[_DEB_LIGN_SAISIE_][0].value = f'01/{mois}/{annee}'

    # Application du report des situations des mois précédents
    if sheet.title[:-5] == 'Janvier':
        sheet[f'{get_c(_NB_COL_INTIT_ + 1)}{_DEB_LIGN_SAISIE_}'].value = 0
        sheet[f'{get_c(_DEB_BANQUE_)}{_DEB_LIGN_SAISIE_}'].value = 0
    else:
        liste_sheet = doc.sheetnames
        indice_sheet = liste_sheet.index(sheet.title)
        sheet[f'{get_c(_NB_COL_INTIT_ + 1)}{_DEB_LIGN_SAISIE_}'].value = \
            f'=\'{liste_sheet[indice_sheet - 1]}\'!{get_c(_DEB_BANQUE_ - 1)}{_DEB_LIGN_SAISIE_ - 2}'
        sheet[f'{get_c(_DEB_BANQUE_)}{_DEB_LIGN_SAISIE_}'].value = \
            f'=\'{liste_sheet[indice_sheet - 1]}\'!{get_c(_DEB_RECETTES_ - 1)}{_DEB_LIGN_SAISIE_ - 2}'


def mettreenformesheetbilan(sheet):
    """
    Met en forme la feuille de bilan comptable de l'année
    :param sheet: Feuille de calcul concernée
    :return:
    """
    _NB_COL_DEPENSES_ = len(_LIB_DEPENSES_)
    _NUM_LAST_COL_ = _DEB_DEPENSES_BILAN_ + _NB_COL_DEPENSES_
    print(f'Dernière colonne : {get_c(_NUM_LAST_COL_)}')
    # Dimensionnement des colonnes
    for i in range(1, _NUM_LAST_COL_):
        sheet.column_dimensions[f'{get_c(i)}'].width = 9
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions[f'{get_c(_NUM_LAST_COL_)}'].width = 12

    # Dimensionnement des lignes
    list_hauteur_ligne = (16, 35, 24.5, 3.5)
    # Tableau général
    for i in range(len(list_hauteur_ligne)):
        sheet.row_dimensions[i + 1].height = list_hauteur_ligne[i]
    for i in range(_DEB_LIGN_SAISIE_B_, _DEB_LIGN_SAISIE_B_ + _NB_MOIS_):
        sheet.row_dimensions[i].height = 12.5
    # Tableau totaux
    list_hauteur_ligne = (15, 20.5, 24.5)
    for i in range(len(list_hauteur_ligne)):
        sheet.row_dimensions[i + _L_TAB_TOTAUX_].height = \
            list_hauteur_ligne[i]

    # Fusion des cellules
    list_cells_a_merge = (f'B1:{get_c(1 + _NB_COL_BANQUE_)}1',
                          f'{get_c(_DEB_DEPENSES_BILAN_)}1:'
                          f'{get_c(_DEB_DEPENSES_BILAN_ + _NB_COL_DEPENSES_ - 1)}1',
                          f'{get_c(_DEB_DEPENSES_BILAN_)}{_L_TAB_TOTAUX_}:'
                          f'{get_c(_DEB_DEPENSES_BILAN_ + _NB_COL_DEPENSES_ - 1)}{_L_TAB_TOTAUX_}')
    for i in range(len(list_cells_a_merge)):
        sheet.merge_cells(list_cells_a_merge[i])

    # Application des alignements spécifiques
    # Tableau principal
    for row in sheet[f'A1:{get_c(_DEB_DEPENSES_BILAN_ + _NB_COL_DEPENSES_)}2']:
        for cell in row:
            cell.alignment = align_center
    sheet[f'{get_c(_NUM_LAST_COL_)}3'].alignment = align_center
    for row in sheet[f'A3:{get_c(_NUM_LAST_COL_)}{_BTM_BILAN_}']:
        for cell in row:
            cell.alignment = align_droite
    # Tableau résumé
    sheet[f'{get_c(_DEB_DEPENSES_BILAN_)}{_L_TAB_TOTAUX_}'].alignment = align_center_adjust
    for row in sheet[(f'{get_c(_DEB_DEPENSES_BILAN_)}{_L_TAB_TOTAUX_ + 1}:'
                      f'{get_c(_DEB_DEPENSES_BILAN_ + _NB_COL_DEPENSES_)}{_L_TAB_TOTAUX_ + 2}')]:
        for cell in row:
            cell.alignment = align_center

    # Application des Fonts
    for row in sheet[f'A2:{get_c(_NUM_LAST_COL_ - 1)}{_L_TAB_TOTAUX_ + 2}']:
        for cell in row:
            cell.font = font_huit
    sheet['B1'].font = font_huit_bold
    sheet[f'{get_c(_DEB_DEPENSES_BILAN_)}1'].font = font_huit_bold
    sheet[f'{get_c(_NUM_LAST_COL_)}2'].font = font_dix_bold
    sheet[f'{get_c(_NUM_LAST_COL_)}3'].font = font_dix
    sheet[f'{get_c(_DEB_DEPENSES_BILAN_)}{_L_TAB_TOTAUX_}'].font = font_dix
    for i in range(_DEB_LIGN_SAISIE_B_, _DEB_LIGN_SAISIE_B_ + _NB_MOIS_ + 1):
        sheet[f'{get_c(_NUM_LAST_COL_)}{i}'].font = font_dix_rouge

    # Application des couleurs de fonds de cellules
    for i in range(_NUM_LAST_COL_):
        sheet[f'{get_c(i + 1)}4'].fill = fill_jaune
    sheet['A3'].fill = fill_gris

    # Formats du contenu des cellules
    monetaire_euro = '#,##0.00 €'
    for row in sheet[f'B3:{get_c(_NUM_LAST_COL_)}{_BTM_BILAN_}']:
        for cell in row:
            cell.number_format = monetaire_euro
    for i in range(_DEB_DEPENSES_BILAN_, _NUM_LAST_COL_):
        sheet[f'{get_c(i)}{_L_TAB_TOTAUX_ + 2}'].number_format = monetaire_euro

    # Définition des zones multi-cellules
    liste_zones = (f'B1:{get_c(_NUM_LAST_COL_ - 1)}{_BTM_BILAN_}',  # Général
                   'A2:A2', f'B3:{get_c(_DEB_DEPENSES_BILAN_ - 1)}{_BTM_BILAN_}',  # Données banque
                   f'{get_c(_NUM_LAST_COL_)}2:{get_c(_NUM_LAST_COL_)}2',  # TOTAL droite
                   f'{get_c(_DEB_DEPENSES_BILAN_)}3:'
                   f'{get_c(_NUM_LAST_COL_ - 1)}{_BTM_BILAN_}',  # Données dépenses
                   f'B1:{get_c(_DEB_DEPENSES_BILAN_ - 1)}1',  # Banque
                   f'{get_c(_DEB_DEPENSES_BILAN_)}1:{get_c(_NUM_LAST_COL_ - 1)}1',  # Titre
                   f'B2:{get_c(_DEB_DEPENSES_BILAN_ - 1)}2',  # Intitulés banque
                   f'{get_c(_DEB_DEPENSES_BILAN_)}2:{get_c(_NUM_LAST_COL_ - 1)}2',  # Intit dép
                   f'A3:A{_BTM_BILAN_}',  # Dates
                   f'{get_c(_NUM_LAST_COL_)}3:{get_c(_NUM_LAST_COL_)}{_BTM_BILAN_}',  # Totaux
                   f'{get_c(_DEB_DEPENSES_BILAN_)}{_L_TAB_TOTAUX_}:'
                   f'{get_c(_NUM_LAST_COL_ - 1)}{_L_TAB_TOTAUX_}',  # Titre tableau résumé
                   f'{get_c(_DEB_DEPENSES_BILAN_)}{_L_TAB_TOTAUX_ + 1}:'
                   f'{get_c(_NUM_LAST_COL_ - 1)}{_L_TAB_TOTAUX_ + 1}',  # Intitulés des dépenses tab. résumé
                   f'{get_c(_DEB_DEPENSES_BILAN_)}{_L_TAB_TOTAUX_ + 2}:'
                   f'{get_c(_NUM_LAST_COL_ - 1)}{_L_TAB_TOTAUX_ + 2}')  # Résumé dépenses par type

    # Application des bordures
    for zone in liste_zones:
        appliquerbordures(sheet[zone])


def remplirsheetbilan(doc):
    """
    Remplit la feuille de calcul du bilan annuel
    :param doc: Document au format .xlxs
    :return:
    """
    _NB_COL_DEPENSES_ = len(_LIB_DEPENSES_)
    _NUM_LAST_COL_ = _DEB_DEPENSES_BILAN_ + _NB_COL_DEPENSES_
    nbligne = int(lignes.get()) + 6
    # Récupère la liste des onglets
    liste_sheet = []
    for sheetname in doc.sheetnames:
        liste_sheet.append(sheetname)

    # Remplit la liste des dépenses dans la feuille Bilan
    sheet = doc[liste_sheet[len(liste_sheet) - 1]]
    for i in range(_NB_COL_DEPENSES_):
        sheet[f'{get_c(_DEB_DEPENSES_BILAN_ + i)}2'].value = _LIB_DEPENSES_[i]
        sheet[f'{get_c(_DEB_DEPENSES_BILAN_ + i)}{_L_TAB_TOTAUX_ + 1}'].value = _LIB_DEPENSES_[i]

    # Remplit les en-têtes
    sheet['B1'].value = 'Banque'
    sheet[f'{get_c(_DEB_DEPENSES_BILAN_)}1'].value = 'Dépenses'
    sheet[f'{get_c(_DEB_DEPENSES_BILAN_)}{_L_TAB_TOTAUX_}'].value = 'Dépenses mensuelles moyennes'
    sheet[f'{get_c(_NUM_LAST_COL_)}2'].value = 'TOTAL'
    list_entete = ('DATE', 'Recettes', 'Dépenses', 'Situation')
    for i in range(len(list_entete)):
        sheet[f'{get_c(i + 1)}2'].value = list_entete[i]

    # Remplit les formules de calcul des sommes et des moyennes des dépenses
    sheet[f'{get_c(_DEB_BANQUE_BILAN_)}3'].value =\
        (f'=SUM({get_c(_DEB_BANQUE_BILAN_)}{_DEB_LIGN_SAISIE_B_}:'
         f'{get_c(_DEB_BANQUE_BILAN_)}{_BTM_BILAN_})')
    sheet[f'{get_c(_DEB_BANQUE_BILAN_ + 1)}3'].value =\
        (f'=SUM({get_c((_DEB_BANQUE_BILAN_ + 1))}{_DEB_LIGN_SAISIE_B_}:'
         f'{get_c((_DEB_BANQUE_BILAN_ + 1))}{_BTM_BILAN_})')
    sheet[f'{get_c(_DEB_DEPENSES_BILAN_ - 1)}3'].value = \
        f'={get_c(_DEB_BANQUE_BILAN_)}3-{get_c(_DEB_BANQUE_BILAN_ + 1)}3'
    for i in range(_DEB_DEPENSES_BILAN_, _NUM_LAST_COL_):
        sheet[f'{get_c(i)}3'].value =\
            f'=SUM({get_c(i)}{_DEB_LIGN_SAISIE_B_}:{get_c(i)}{_BTM_BILAN_})'
        sheet[f'{get_c(i)}{_L_TAB_TOTAUX_ + 2}'].value =\
            f'=AVERAGE({get_c(i)}{_DEB_LIGN_SAISIE_B_}:{get_c(i)}{_BTM_BILAN_})'

    # Remplit la cellule de total des dépenses
    sheet[f'{get_c(_NUM_LAST_COL_)}3'].value =\
        f'=SUM({get_c(_DEB_DEPENSES_BILAN_)}3:{get_c(_NUM_LAST_COL_ - 1)}3)'

    # Remplit la colonne des totaux
    for i in range(_DEB_LIGN_SAISIE_B_, _DEB_LIGN_SAISIE_B_ + _NB_MOIS_):
        sheet[f'{get_c(_NUM_LAST_COL_)}{i}'].value = \
            (f'={get_c(_DEB_BANQUE_BILAN_ + 1)}{i}-SUM({get_c(_DEB_DEPENSES_BILAN_)}{i}:'
             f'{get_c(_NUM_LAST_COL_ - 1)}{i})')

    # Remplit la colonne des mois
    for i in range(_NB_MOIS_):
        sheet[f'A{_DEB_LIGN_SAISIE_B_ + i}'].value = liste_sheet[i]

    # Remplit le contenu du tableau principal
    # Partie "Banque"
    c_banq_b = get_c(_DEB_BANQUE_BILAN_)
    l_saisie = _DEB_LIGN_SAISIE_ + 1
    deb_caisse = _NB_COL_INTIT_ + 1
    bottom = nbligne - 1
    for i in range(_DEB_LIGN_SAISIE_B_, _DEB_LIGN_SAISIE_B_ + _NB_MOIS_):
        feuille = liste_sheet[i - _DEB_LIGN_SAISIE_B_]
        sheet[f'{c_banq_b}{i}'].value = \
            f'=SUM(\'{feuille}\'!{get_c(deb_caisse)}{l_saisie}:\'{feuille}\'!{get_c(deb_caisse)}{bottom})+' \
            f'SUM(\'{feuille}\'!{get_c(_DEB_BANQUE_)}{l_saisie}:\'{feuille}\'!{get_c(_DEB_BANQUE_)}{bottom})'
        sheet[f'{get_c(_DEB_BANQUE_BILAN_ + 1)}{i}'].value = \
            f'=SUM(\'{feuille}\'!{get_c(deb_caisse + 1)}{l_saisie}:\'{feuille}\'!{get_c(deb_caisse + 1)}{bottom})+' \
            f'SUM(\'{feuille}\'!{get_c(_DEB_BANQUE_ + 1)}{l_saisie}:\'{feuille}\'!{get_c(_DEB_BANQUE_ + 1)}{bottom})'
        sheet[f'D{i}'].value = f'={c_banq_b}{i}-{get_c(_DEB_BANQUE_BILAN_ + 1)}{i}'
        # Partie "Dépenses"
        for j in range(_NB_COL_DEPENSES_):
            sheet[f'{get_c(_DEB_DEPENSES_BILAN_ + j)}{i}'].value =\
                f'=\'{feuille}\'!{get_c(_DEB_DEPENSES_ + j)}{_DEB_LIGN_SAISIE_ - 2}'


def cheminfichier():
    """
    Définit le chemin complet du fichier (répertoire de destination et nom de fichier)
    """
    # Chemin vers le répertoire de destination
    path = f'{askdirectory(title="Choix du dossier de destination")}'
    # Attribution du nom de fichier
    lblpath['text'] = f'{path}'


def appliquerbordures(zone):
    """
    Applique les bordures à une zone de feuille de calcul (medium aux bordures extérieures et thin aux autres)
    :param zone: Zone sur laquelle porte l'application des bordures
    :return:
    """
    # Nombre de lignes dans la zone
    nligne = len(zone)
    # Nombre de colonnes dans la zone
    ncolonne = len(zone[0])
    idrow = 0
    # Si la zone ne s'étend que sur une seule ligne
    if nligne == 1:
        # Si la zone ne comporte qu'une cellule
        if ncolonne == 1:
            zone[0][0].border = medium_all_borders
            return
        for row in zone:
            idcell = 0
            # Parcourt chaque cellule de la ligne
            for cell in row:
                # Si la cellule est la première de la ligne
                if idcell == 0:
                    cell.border = medium_contour_left
                # Si la cellule est la dernière de la ligne
                elif idcell == len(row) - 1:
                    cell.border = medium_contour_right
                # Si la cellule n'est ni la première ni la dernière de la ligne
                else:
                    cell.border = medium_top_bottom
                idcell += 1
    # Si la zone s'étend sur plusieurs lignes
    else:
        # Parcourt chaque ligne de la zone
        for row in zone:
            idcell = 0
            # Si la zone ne s'étend que sur une seule colonne
            if len(row) == 1:
                # Parcourt chaque cellule de la colonne
                for cell in row:
                    # Si la cellule est la première de la colonne
                    if idrow == 0:
                        cell.border = medium_contour_top
                    # Si la cellule est la dernière de la colonne
                    elif idrow == nligne - 1:
                        cell.border = medium_contour_bottom
                    # Si la cellule n'est ni la première ni la dernière de la colonne
                    else:
                        cell.border = medium_left_right
            # Si la zone comporte plusieurs lignes ET plusieurs colonnes
            else:
                # Parcourt chaque cellule de la ligne en cours
                for cell in row:
                    # Si la ligne est la première de la zone
                    if idrow == 0:
                        # Si la cellule est la première de la ligne
                        if idcell == 0:
                            cell.border = medium_coin_top_left
                        # Si la cellule est la dernière de la ligne
                        elif idcell == len(row) - 1:
                            cell.border = medium_coin_top_right
                        else:
                            cell.border = medium_border_top
                    # Si la ligne est la dernière de la zone
                    elif idrow == nligne - 1:
                        # Si la cellule est la première de la ligne
                        if idcell == 0:
                            cell.border = medium_coin_bottom_left
                        # Si la cellule est la dernière de la ligne
                        elif idcell == len(row) - 1:
                            cell.border = medium_coin_bottom_right
                        else:
                            cell.border = medium_border_bottom
                    # Si la ligne n'est ni la première ni la dernière de la zone
                    else:
                        # Si la cellule est la première de la ligne
                        if idcell == 0:
                            cell.border = medium_border_left
                        # Si la cellule est la dernière de la ligne
                        elif idcell == len(row) - 1:
                            cell.border = medium_border_right
                        else:
                            cell.border = thin_all_borders
                    idcell += 1
            idrow += 1


def verifcontenu():
    """
    Vérifie si le contenu des champs de la fenêtre sont corrects
    :return: True si ok, False dans le cas contraire
    """
    # Vérifie si l'année renseignée est correcte
    if lannee.get() == '':
        lblerror['text'] = "Vous devez saisir une année"
        txtannee.focus()
        return False
    else:
        try:
            year = int(lannee.get())
            if not 999 < year < 10000:
                lblerror['text'] = "L'année doit être comprise entre 1000 et 9999"
                return False
        except ValueError:
            lblerror['text'] = "L'année doit être un nombre entier"
            return False
        # Vérifie si le nombre de lignes de saisie renseigné est correct
        if lignes.get() == '':
            lblerror['text'] = "Vous devez entrer un nombre de lignes de saisie"
            txtnblign.focus()
            return False
        else:
            try:
                ln = int(lignes.get())
                if not 9 < ln < 1000:
                    lblerror['text'] = "Le nombre de lignes doit être compris entre 10 et 999"
                    return False
            except ValueError:
                lblerror['text'] = "Le nombre de lignes doit être un nombre entier"
                return False
            # Vérifie si un nom de fichier a été saisi
            if lenom.get() == '':
                lblerror['text'] = "Vous devez entrer un nom de fichier"
                txtnom.focus()
                return False
            # Vérifie si un répertoire de destination a été choisi
            elif lblpath['text'] == '':
                lblerror['text'] = "Vous devez sélectionner un dossier de destination"
                btndir.focus()
                return False
            if len(_LIB_DEPENSES_) < 2:
                lblerror['text'] = "Vous devez saisir au moins deux types de dépenses"
                txtdep.focus()
                return False
            else:
                lblerror['text'] = ''
                return True


def creerfichier():
    """
    Génère le fichier de comptabilité
    :return:
    """
    if verifcontenu():
        document = genererfichiercompta()
        # Crée le chemin complet pour l'enregistrement du fichier
        nom = verifnom()
        dest = lblpath['text']
        path = f'{dest}/{nom}.xlsx'
        # Enregistre le document dans le répertoire choisi
        document.save(path)
        fenetre.destroy()


def adddepense():
    global _LIB_DEPENSES_
    if len(_LIB_DEPENSES_) >= 15:
        lblerror['text'] = "Nombre maximum de type de dépenses atteint"
        return
    if txtdep.get() != '':
        if txtdep.get() in _LIB_DEPENSES_:
            lblerror['text'] = "Cette catégorie est déjà présente dans la liste"
            return
        else:
            _LIB_DEPENSES_.append(txtdep.get())
            txtdep['text'] = ''
            affichedepenses()
    else:
        lblerror['text'] = "Veuillez saisir une catégorie de dépense pour l'ajouter"


def suppdepense():
    try:
        del _LIB_DEPENSES_[len(_LIB_DEPENSES_) - 1]
        affichedepenses()
    except IndexError:
        lblerror['text'] = "La liste des dépenses est vide"


def affichedepenses():
    txtdep.delete(0, len(txtdep.get()))
    lbllistdep['text'] = ''
    for i in range(len(_LIB_DEPENSES_)):
        if i != 0:
            lbllistdep['text'] += ' | '
        lbllistdep['text'] += f'{_LIB_DEPENSES_[i]}'


# Création de l'interface graphique
fenetre = Tk()
lenom = tkinter.StringVar(name='nomfic', master=fenetre)
lannee = tkinter.StringVar(name='annee', master=fenetre)
lignes = tkinter.StringVar(name='nbligne', master=fenetre)
largeur = 380
fenetre.geometry(f'{largeur}x500')
fenetre.title("Création d'un fichier de comptabilité")
fenetre.resizable(width=False, height=False)

# Contenu de l'interface graphique
# Couleur de fond des Entry et Label
bg_entry = '#E4FFE7'
arial_dix = ('Arial', 10)
# Titre
lblprs = Label(fenetre, text="Informations du document", font=('Arial', 14, "bold"))
lblprs.pack(pady=10)

# Zone du choix de l'année
fenannee = Frame(fenetre)
fenannee.pack(pady=6)
lblannee = Label(fenannee, text="Année comptable :", font=arial_dix)
lblannee.pack(side=LEFT)
txtannee = Entry(fenannee, border=2, width=7, justify='center', font=arial_dix, textvariable=lannee,
                 bg=bg_entry)
txtannee.pack(side=LEFT)

# Zone du choix de nb de lignes de saisie
fennblign = Frame(fenetre)
fennblign.pack(pady=6)
lblnblign = Label(fennblign, text="Nombre de lignes de saisie :", font=arial_dix)
lblnblign.pack(side=LEFT)
txtnblign = Entry(fennblign, border=2, width=5, justify='center', font=arial_dix, textvariable=lignes,
                  bg=bg_entry)
txtnblign.pack(side=LEFT)

# Zone du choix de nom
fennom = Frame(fenetre)
fennom.pack(pady=6)
lblnom = Label(fennom, text="Nom du fichier :", font=arial_dix)
lblnom.pack(side=LEFT)
txtnom = Entry(fennom, border=2, width=25, font=arial_dix, textvariable=lenom, bg=bg_entry)
txtnom.pack(side=LEFT)

# Zone du choix de répertoire
fendir = Frame(fenetre)
fendir.pack()
lbldir = Label(fendir, text="Destination :", font=arial_dix)
lbldir.pack(side=LEFT)
btndir = Button(fendir, text="...", width=3, font=arial_dix, command=cheminfichier, bg=bg_entry)
btndir.pack(side=LEFT, padx=2)
lblpath = Label(fenetre, width=25, text='', font=arial_dix, bg=bg_entry, anchor="w", wraplength=largeur)
lblpath.pack(side=TOP, pady=2, padx=10, fill=X)
lblspace = Label(fenetre, text='')
lblspace.pack(side=TOP)

# Zone de la liste des dépenses
lblintrolist = Label(fenetre, text='Liste des catégories de dépenses', font=arial_dix, anchor="w")
lblintrolist.pack(side=TOP, padx=10, fill=X)
fendep = Frame(fenetre)
fendep.pack(side=TOP, pady=6)
ladepense = tkinter.StringVar(name='ladepense', master=fendep)
lbllistdep = Label(fendep, font=arial_dix, bg=bg_entry, height=3, width=largeur, wraplength=largeur - 20, anchor="w")
lbllistdep.pack(side=TOP, padx=10, fill=X)
txtdep = Entry(fendep, font=arial_dix, textvariable=ladepense, bg=bg_entry)
txtdep.pack(side=BOTTOM, pady=7)

# Zone des boutons d'ajout et de suppression
zonebtngestliste = Frame(fenetre)
zonebtngestliste.pack(pady=4, side=TOP)
btnadddep = Button(zonebtngestliste, text='Ajouter', bg='green', fg='white', font=('Arial', 10, 'bold'),
                   command=adddepense)
btnadddep.pack(pady=3, side=LEFT, padx=4)
btnadddep = Button(zonebtngestliste, text='Supprimer', bg='red', fg='white', font=('Arial', 10, 'bold'),
                   command=suppdepense)
btnadddep.pack(pady=3, side=LEFT, padx=4)

# Bouton de création du document
btnsend = Button(fenetre, text="Lancer la génération", font=('Arial', 12, 'bold'), bg='darkgreen', fg='white',
                 command=creerfichier)
btnsend.pack(pady=18)

# Label des messages d'erreur
lblerror = Label(fenetre, text='', font=arial_dix, fg='red')
lblerror.pack()

affichedepenses()
fenetre.mainloop()
